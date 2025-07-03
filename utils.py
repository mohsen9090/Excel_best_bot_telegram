#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import shutil
import logging

from config import *

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def ensure_directories():
    directories = [EXCEL_DIR, FIELDS_DIR, USER_THEMES_DIR, BACKUP_DIR]
    for directory in directories:
        if not os.path.exists(directory):
            os.makedirs(directory)
            logger.info(f"پوشه {directory} ایجاد شد")

def ensure_excel_file(chat_id):
    ensure_directories()
    excel_file = get_excel_file(chat_id)
    
    if not os.path.exists(excel_file):
        fields = load_fields(chat_id)
        df = pd.DataFrame(columns=fields)
        theme = load_user_theme(chat_id)
        create_excel(df, theme, chat_id)
        logger.info(f"فایل Excel جدید برای chat_id {chat_id} ایجاد شد")

def load_fields(chat_id):
    ensure_directories()
    fields_file = get_fields_file(chat_id)
    
    if os.path.exists(fields_file):
        try:
            with open(fields_file, 'r', encoding='utf-8') as f:
                fields = json.load(f)
                return fields if fields else DEFAULT_FIELDS.copy()
        except Exception as e:
            logger.error(f"خطا در بارگذاری فیلدهای chat_id {chat_id}: {e}")
    
    save_fields(DEFAULT_FIELDS.copy(), chat_id)
    return DEFAULT_FIELDS.copy()

def save_fields(fields, chat_id):
    ensure_directories()
    fields_file = get_fields_file(chat_id)
    
    try:
        with open(fields_file, 'w', encoding='utf-8') as f:
            json.dump(fields, f, ensure_ascii=False, indent=2)
        logger.info(f"فیلدهای chat_id {chat_id} ذخیره شد")
        return True
    except Exception as e:
        logger.error(f"خطا در ذخیره فیلدهای chat_id {chat_id}: {e}")
        return False

def load_user_theme(chat_id):
    ensure_directories()
    theme_file = get_user_theme_file(chat_id)
    
    if os.path.exists(theme_file):
        try:
            with open(theme_file, 'r', encoding='utf-8') as f:
                theme_data = json.load(f)
                theme = theme_data.get('theme', 'blue')
                return theme if theme in THEMES else 'blue'
        except Exception as e:
            logger.error(f"خطا در بارگذاری تم chat_id {chat_id}: {e}")
    
    save_user_theme(chat_id, 'blue')
    return 'blue'

def save_user_theme(chat_id, theme):
    ensure_directories()
    theme_file = get_user_theme_file(chat_id)
    
    try:
        theme_data = {
            'theme': theme,
            'updated_at': datetime.now().isoformat(),
            'chat_id': chat_id
        }
        
        with open(theme_file, 'w', encoding='utf-8') as f:
            json.dump(theme_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"تم {theme} برای chat_id {chat_id} ذخیره شد")
        return True
    except Exception as e:
        logger.error(f"خطا در ذخیره تم chat_id {chat_id}: {e}")
        return False

def format_numeric_field(value, field_name):
    """فرمت کردن فیلدهای عددی برای جلوگیری از نمایش .0"""
    if pd.isna(value) or value is None:
        return ""
    
    str_value = str(value).strip()
    
    if not str_value:
        return ""
    
    # فیلدهای خاص که باید به صورت رشته ذخیره شوند
    field_lower = field_name.lower()
    is_special_field = any(special in field_lower for special in 
                          ['کد ملی', 'کارت بانک', 'تلفن', 'موبایل', 'پست', 'شناسنامه'])
    
    if is_special_field:
        # حذف .0 از انتها اگر وجود دارد
        if str_value.endswith('.0'):
            str_value = str_value[:-2]
        
        # فقط اعداد را نگه دار
        clean_value = ''.join(c for c in str_value if c.isdigit())
        return clean_value
    
    # برای سایر فیلدها
    try:
        if '.' in str_value and str_value.endswith('.0'):
            return str_value[:-2]
        return str_value
    except:
        return str_value

def create_excel(df, theme, chat_id):
    try:
        ensure_directories()
        excel_file = get_excel_file(chat_id)
        
        if df.empty:
            fields = load_fields(chat_id)
            df = pd.DataFrame(columns=fields)
        
        # پردازش داده‌ها برای جلوگیری از مشکل .0
        processed_df = df.copy()
        for col in processed_df.columns:
            processed_df[col] = processed_df[col].apply(lambda x: format_numeric_field(x, col))
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"داده‌های کاربر {chat_id}"
        
        theme_config = THEMES.get(theme, THEMES['blue'])
        header_fill = PatternFill(start_color=theme_config['header'], 
                                 end_color=theme_config['header'], 
                                 fill_type='solid')
        row1_fill = PatternFill(start_color=theme_config['row1'], 
                               end_color=theme_config['row1'], 
                               fill_type='solid')
        row2_fill = PatternFill(start_color=theme_config['row2'], 
                               end_color=theme_config['row2'], 
                               fill_type='solid')
        
        header_font = Font(bold=True, color='FFFFFF', size=12)
        normal_font = Font(size=11)
        border = Border(
            left=Side(style='thin', color=theme_config.get('border', '000000')),
            right=Side(style='thin', color=theme_config.get('border', '000000')),
            top=Side(style='thin', color=theme_config.get('border', '000000')),
            bottom=Side(style='thin', color=theme_config.get('border', '000000'))
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # اضافه کردن هدرها
        for col_idx, column in enumerate(processed_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment
            
            # تنظیم عرض ستون بر اساس نوع فیلد
            column_letter = cell.column_letter
            if any(special in column.lower() for special in ['کارت بانک', 'شماره کارت']):
                ws.column_dimensions[column_letter].width = 20
            elif any(special in column.lower() for special in ['کد ملی', 'تلفن', 'موبایل']):
                ws.column_dimensions[column_letter].width = 15
            elif 'آدرس' in column.lower():
                ws.column_dimensions[column_letter].width = 25
            else:
                ws.column_dimensions[column_letter].width = 12
        
        # اضافه کردن داده‌ها
        for row_idx, (_, row) in enumerate(processed_df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                display_value = str(value) if pd.notna(value) and value != "" else ""
                
                cell = ws.cell(row=row_idx, column=col_idx, value=display_value)
                cell.font = normal_font
                cell.border = border
                cell.alignment = center_alignment
                
                if row_idx % 2 == 0:
                    cell.fill = row1_fill
                else:
                    cell.fill = row2_fill
        
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        
        footer_row = len(processed_df) + 3
        ws.cell(row=footer_row, column=1, 
               value=f"تولید شده توسط ربات Excel - {datetime.now().strftime('%Y/%m/%d %H:%M')}")
        ws.cell(row=footer_row + 1, column=1, 
               value=f"Chat ID: {chat_id} | تم: {theme_config['name']}")
        
        wb.save(excel_file)
        logger.info(f"فایل Excel برای chat_id {chat_id} با تم {theme} ایجاد شد")
        
        if AUTO_BACKUP and len(processed_df) > 0:
            create_backup(chat_id)
        
        return True
        
    except Exception as e:
        logger.error(f"خطا در ایجاد فایل Excel برای chat_id {chat_id}: {e}")
        return False

def create_backup(chat_id):
    try:
        ensure_directories()
        excel_file = get_excel_file(chat_id)
        
        if not os.path.exists(excel_file):
            return False
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = get_backup_file(chat_id, timestamp)
        
        shutil.copy2(excel_file, backup_file)
        cleanup_old_backups(chat_id)
        
        logger.info(f"بکاپ برای chat_id {chat_id} ایجاد شد: {backup_file}")
        return True
        
    except Exception as e:
        logger.error(f"خطا در ایجاد بکاپ برای chat_id {chat_id}: {e}")
        return False

def cleanup_old_backups(chat_id):
    try:
        backup_pattern = f"backup_{chat_id}_"
        backup_files = []
        
        for filename in os.listdir(BACKUP_DIR):
            if filename.startswith(backup_pattern) and filename.endswith('.xlsx'):
                backup_files.append(os.path.join(BACKUP_DIR, filename))
        
        backup_files.sort(key=os.path.getmtime, reverse=True)
        
        if len(backup_files) > MAX_BACKUPS_PER_USER:
            for old_backup in backup_files[MAX_BACKUPS_PER_USER:]:
                os.remove(old_backup)
                logger.info(f"بکاپ قدیمی حذف شد: {old_backup}")
                
    except Exception as e:
        logger.error(f"خطا در پاک‌سازی بکاپ‌های chat_id {chat_id}: {e}")

def validate_field_input(field, value):
    if not ENABLE_FIELD_VALIDATION:
        return True, value
    
    value = str(value).strip()
    
    if len(value) > MAX_FIELD_LENGTH:
        return False, f"❌  طول {field} نباید از {MAX_FIELD_LENGTH} کاراکتر بیشتر باشد."
    
    field_lower = field.lower()
    
    if 'کد ملی' in field or 'ملی' in field_lower:
        if value.endswith('.0'):
            value = value[:-2]
        clean_value = ''.join(c for c in value if c.isdigit())
        if len(clean_value) != 10:
            return False, f"❌  {field} باید دقیقاً 10 رقم باشد."
        value = clean_value
    
    elif 'تلفن' in field_lower or 'موبایل' in field_lower:
        if value.endswith('.0'):
            value = value[:-2]
        clean_phone = ''.join(c for c in value if c.isdigit())
        if len(clean_phone) < 10:
            return False, f"❌  {field} نامعتبر است."
        value = clean_phone
    
    elif 'کارت بانک' in field_lower or 'شماره کارت' in field_lower:
        if value.endswith('.0'):
            value = value[:-2]
        clean_card = ''.join(c for c in value if c.isdigit())
        if len(clean_card) not in [16, 19]:
            return False, f"❌  {field} باید 16 رقم باشد."
        value = clean_card
    
    elif 'کد پستی' in field_lower:
        if value.endswith('.0'):
            value = value[:-2]
        clean_postal = ''.join(c for c in value if c.isdigit())
        if len(clean_postal) != 10:
            return False, f"❌  {field} باید 10 رقم باشد."
        value = clean_postal
    
    elif 'ایمیل' in field_lower or 'email' in field_lower:
        if '@' not in value or '.' not in value:
            return False, f"❌  {field} نامعتبر است."
    
    elif 'سن' in field_lower or 'age' in field_lower:
        try:
            if value.endswith('.0'):
                value = value[:-2]
            age = int(value)
            if age < 0 or age > 150:
                return False, f"❌  {field} باید بین 0 تا 150 باشد."
            value = str(age)
        except ValueError:
            return False, f"❌  {field} باید عدد باشد."
    
    elif 'شناسنامه' in field_lower:
        if value.endswith('.0'):
            value = value[:-2]
        clean_id = ''.join(c for c in value if c.isdigit())
        value = clean_id
    
    return True, value

def clean_value(value):
    if pd.isna(value) or value is None:
        return ""
    
    str_value = str(value).strip()
    
    if str_value.endswith('.0'):
        str_value = str_value[:-2]
    
    return str_value

def format_record_display(df, max_records, chat_id=None):
    if df.empty:
        return f"📭  هیچ رکوردی در فایل Chat ID {chat_id} وجود ندارد."
    
    total_records = len(df)
    display_count = min(max_records, total_records)
    
    message = f"📋  **نمایش رکوردهای Chat ID {chat_id}**\n"
    message += f"📊  تعداد کل: {total_records:,} رکورد\n"
    message += f"👁️ نمایش: {display_count} رکورد اول\n\n"
    
    for i in range(display_count):
        record_info = []
        row = df.iloc[i]
        
        for col in df.columns[:4]:
            value = clean_value(row[col])
            if value:
                record_info.append(f"{col}: {value}")
        
        message += f"**{i+1}.** {' | '.join(record_info)}\n"
    
    if total_records > display_count:
        message += f"\n... و {total_records - display_count:,} رکورد دیگر"
        message += f"\n\n💡  برای مشاهده همه از 'دریافت فایل' استفاده کنید."
    
    return message

def get_user_statistics(chat_id):
    try:
        stats = {
            'chat_id': chat_id,
            'total_records': 0,
            'total_fields': 0,
            'file_size': 0,
            'theme': 'نامشخص',
            'last_updated': 'نامشخص',
            'backup_count': 0
        }
        
        excel_file = get_excel_file(chat_id)
        if os.path.exists(excel_file):
            df = pd.read_excel(excel_file)
            stats['total_records'] = len(df)
            stats['total_fields'] = len(df.columns)
            stats['file_size'] = round(os.path.getsize(excel_file) / 1024, 1)
            stats['last_updated'] = datetime.fromtimestamp(
                os.path.getmtime(excel_file)
            ).strftime('%Y/%m/%d %H:%M')
        
        theme = load_user_theme(chat_id)
        stats['theme'] = THEMES[theme]['name']
        
        backup_pattern = f"backup_{chat_id}_"
        backup_count = 0
        if os.path.exists(BACKUP_DIR):
            for filename in os.listdir(BACKUP_DIR):
                if filename.startswith(backup_pattern):
                    backup_count += 1
        stats['backup_count'] = backup_count
        
        return stats
        
    except Exception as e:
        logger.error(f"خطا در دریافت آمار chat_id {chat_id}: {e}")
        return None

def search_in_dataframe(df, query, chat_id):
    if df.empty:
        return []
    
    query = query.lower().strip()
    found_records = []
    
    for index, row in df.iterrows():
        for col in df.columns:
            cell_value = clean_value(row[col]).lower()
            if query in cell_value:
                found_records.append({
                    'index': index + 1,
                    'row_data': row,
                    'matched_field': col,
                    'chat_id': chat_id
                })
                break
    
    return found_records[:MAX_SEARCH_RESULTS]

def validate_excel_file(file_path):
    try:
        if not os.path.exists(file_path):
            return False, "فایل یافت نشد"
        
        file_size = os.path.getsize(file_path)
        if file_size > MAX_FILE_SIZE:
            return False, f"حجم فایل نباید از {MAX_FILE_SIZE // (1024*1024)} مگابایت بیشتر باشد"
        
        df = pd.read_excel(file_path)
        
        if len(df) > MAX_RECORDS_PER_USER:
            return False, f"تعداد رکوردها نباید از {MAX_RECORDS_PER_USER:,} بیشتر باشد"
        
        if len(df.columns) > MAX_FIELDS_PER_USER:
            return False, f"تعداد ستون‌ها نباید از {MAX_FIELDS_PER_USER} بیشتر باشد"
        
        return True, "فایل معتبر است"
        
    except Exception as e:
        return False, f"خطا در خواندن فایل: {str(e)}"

def merge_dataframes(existing_df, new_df):
    try:
        for df in [existing_df, new_df]:
            for col in df.columns:
                df[col] = df[col].apply(lambda x: format_numeric_field(x, col))
        
        existing_cols = set(existing_df.columns)
        new_cols = set(new_df.columns)
        
        if existing_cols != new_cols:
            all_cols = list(existing_cols.union(new_cols))
            
            for col in all_cols:
                if col not in existing_df.columns:
                    existing_df[col] = ""
                if col not in new_df.columns:
                    new_df[col] = ""
            
            existing_df = existing_df[all_cols]
            new_df = new_df[all_cols]
        
        merged_df = pd.concat([existing_df, new_df], ignore_index=True)
        return merged_df
        
    except Exception as e:
        logger.error(f"خطا در ادغام DataFrame: {e}")
        return existing_df

def initialize_user_data(chat_id):
    try:
        ensure_directories()
        save_fields(DEFAULT_FIELDS.copy(), chat_id)
        save_user_theme(chat_id, 'blue')
        ensure_excel_file(chat_id)
        
        logger.info(f"داده‌های اولیه برای chat_id {chat_id} ایجاد شد")
        return True
        
    except Exception as e:
        logger.error(f"خطا در مقداردهی اولیه chat_id {chat_id}: {e}")
        return False
